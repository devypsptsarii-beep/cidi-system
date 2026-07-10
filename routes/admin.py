from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from extensions import db
from models import User, IndustryProfile, ParticipantProfile
from models import TrainingProgram, TrainingRegistration, Certificate, WorkforceDemand
from datetime import date, datetime
from werkzeug.security import generate_password_hash
import uuid
import io
import os

admin = Blueprint('admin', __name__)

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def send_email(subject, recipients, html_body):
    """Helper function to send email safely"""
    try:
        from flask_mail import Message
        from app import mail
        msg = Message(subject=subject, recipients=recipients)
        msg.html = html_body
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

def email_template(title, greeting, body_html, button_text=None, button_url=None):
    """Reusable email HTML template"""
    button_html = ''
    if button_text and button_url:
        button_html = f'''
        <div style="text-align:center;margin:2rem 0;">
            <a href="{button_url}"
               style="background:#1a56db;color:white;padding:12px 28px;
                      border-radius:8px;text-decoration:none;font-weight:bold;">
                {button_text}
            </a>
        </div>'''
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
        <div style="background:#1a56db;padding:2rem;text-align:center;">
            <h1 style="color:white;margin:0;">CIDI 4.0</h1>
            <p style="color:rgba(255,255,255,0.9);margin:0;">Center for Digital Industry 4.0</p>
        </div>
        <div style="padding:2rem;background:#fff;">
            <h2 style="color:#111827;">{title}</h2>
            <p>{greeting}</p>
            {body_html}
            {button_html}
            <p style="color:#6b7280;font-size:0.875rem;margin-top:2rem;">
                If you have questions, contact us at info@cidi40.com
            </p>
        </div>
        <div style="background:#f3f4f6;padding:1rem;text-align:center;
                    color:#6b7280;font-size:0.8rem;">
            © 2026 Center for Digital Industry 4.0. All rights reserved.
        </div>
    </div>
    """

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@admin.route('/dashboard')
@login_required
@admin_required
def dashboard():
    total_participants    = ParticipantProfile.query.count()
    total_industries      = IndustryProfile.query.count()
    total_programs        = TrainingProgram.query.count()
    total_certified       = Certificate.query.count()
    pending_industries    = User.query.filter_by(role='industry', is_approved=False).count()
    pending_registrations = TrainingRegistration.query.filter_by(status='pending').count()
    recent_demands        = WorkforceDemand.query.order_by(
                                WorkforceDemand.submitted_at.desc()).limit(5).all()
    from sqlalchemy import func
    skill_data = db.session.query(
        Certificate.skill_certified,
        func.count(Certificate.id).label('count')
    ).group_by(Certificate.skill_certified).all()
    skill_labels = [s.skill_certified for s in skill_data]
    skill_counts = [s.count           for s in skill_data]

    employed   = ParticipantProfile.query.filter_by(working_status='employed').count()
    unemployed = ParticipantProfile.query.filter_by(working_status='unemployed').count()

    monthly_data = db.session.query(
        func.extract('month', Certificate.issued_date).label('month'),
        func.count(Certificate.id).label('count')
    ).group_by('month').order_by('month').all()
    monthly_labels = ['Jan','Feb','Mar','Apr','May','Jun',
                      'Jul','Aug','Sep','Oct','Nov','Dec']
    monthly_counts = [0] * 12
    for m in monthly_data:
        monthly_counts[int(m.month) - 1] = m.count

    return render_template('admin/dashboard.html',
        total_participants    = total_participants,
        total_industries      = total_industries,
        total_programs        = total_programs,
        total_certified       = total_certified,
        pending_industries    = pending_industries,
        pending_registrations = pending_registrations,
        recent_demands        = recent_demands,
        skill_labels          = skill_labels,
        skill_counts          = skill_counts,
        employed              = employed,
        unemployed            = unemployed,
        monthly_labels        = monthly_labels,
        monthly_counts        = monthly_counts
    )

# ── TRAINING PROGRAMS ─────────────────────────────────────────────────────────
@admin.route('/programs')
@login_required
@admin_required
def programs():
    all_programs = TrainingProgram.query.order_by(
                       TrainingProgram.created_at.desc()).all()
    return render_template('admin/programs.html', programs=all_programs)

@admin.route('/programs/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_program():
    if request.method == 'POST':
        program = TrainingProgram(
            title          = request.form.get('title'),
            description    = request.form.get('description'),
            skill_category = request.form.get('skill_category'),
            start_date     = date.fromisoformat(request.form.get('start_date')),
            end_date       = date.fromisoformat(request.form.get('end_date')),
            capacity       = int(request.form.get('capacity')),
            status         = request.form.get('status', 'open')
        )
        db.session.add(program)
        db.session.commit()
        flash('Training program added successfully!', 'success')
        return redirect(url_for('admin.programs'))
    return render_template('admin/add_program.html')

@admin.route('/programs/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_program(id):
    program = TrainingProgram.query.get_or_404(id)
    if request.method == 'POST':
        program.title          = request.form.get('title')
        program.description    = request.form.get('description')
        program.skill_category = request.form.get('skill_category')
        program.start_date     = date.fromisoformat(request.form.get('start_date'))
        program.end_date       = date.fromisoformat(request.form.get('end_date'))
        program.capacity       = int(request.form.get('capacity'))
        program.status         = request.form.get('status')
        db.session.commit()
        flash('Program updated successfully!', 'success')
        return redirect(url_for('admin.programs'))
    return render_template('admin/edit_program.html', program=program)

@admin.route('/programs/delete/<int:id>')
@login_required
@admin_required
def delete_program(id):
    program = TrainingProgram.query.get_or_404(id)
    db.session.delete(program)
    db.session.commit()
    flash('Program deleted.', 'success')
    return redirect(url_for('admin.programs'))

# ── PARTICIPANTS ──────────────────────────────────────────────────────────────
@admin.route('/participants')
@login_required
@admin_required
def participants():
    search        = request.args.get('search', '')
    city_filter   = request.args.get('city', '')
    status_filter = request.args.get('status', '')
    query         = ParticipantProfile.query
    if search:
        query = query.filter(ParticipantProfile.full_name.ilike(f'%{search}%'))
    if city_filter:
        query = query.filter(ParticipantProfile.city.ilike(f'%{city_filter}%'))
    if status_filter:
        query = query.filter(ParticipantProfile.working_status == status_filter)
    all_participants = query.order_by(ParticipantProfile.full_name).all()
    return render_template('admin/participants.html',
        participants  = all_participants,
        search        = search,
        city_filter   = city_filter,
        status_filter = status_filter
    )

@admin.route('/participants/delete/<int:id>')
@login_required
@admin_required
def delete_participant(id):
    profile = ParticipantProfile.query.get_or_404(id)
    user    = User.query.get_or_404(profile.user_id)

    # Delete related records first
    Certificate.query.filter_by(participant_id=profile.id).delete()
    TrainingRegistration.query.filter_by(participant_id=profile.id).delete()
    db.session.delete(profile)
    db.session.delete(user)
    db.session.commit()
    flash('Participant deleted successfully.', 'success')
    return redirect(url_for('admin.participants'))

@admin.route('/participants/<int:id>')
@login_required
@admin_required
def view_participant(id):
    profile = ParticipantProfile.query.get_or_404(id)
    certs   = Certificate.query.filter_by(participant_id=id).all()
    regs    = TrainingRegistration.query.filter_by(participant_id=id).all()
    return render_template('admin/view_participant.html',
        profile=profile, certs=certs, regs=regs)

# ── EXPORT PARTICIPANTS TO EXCEL ──────────────────────────────────────────────
@admin.route('/export/participants')
@login_required
@admin_required
def export_participants():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Participants'

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='1A56DB',
                               end_color='1A56DB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['No','Full Name','Email','Gender','Place of Birth',
               'Date of Birth','Age','City','Address','Phone',
               'Education','Working Status','Certificates','Source','Registered']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    widths = [5,25,30,10,20,15,8,15,30,18,15,15,12,12,15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    participants = ParticipantProfile.query.order_by(
                       ParticipantProfile.full_name).all()
    for idx, p in enumerate(participants, 1):
        cert_count = Certificate.query.filter_by(participant_id=p.id).count()
        source     = 'Imported' if p.user.is_imported else 'Self-registered'
        row = [idx, p.full_name, p.user.email, p.gender,
               p.place_of_birth, p.date_of_birth.strftime('%d-%m-%Y'),
               p.age, p.city or '-', p.address, p.phone_number,
               p.education, p.working_status or 'unemployed',
               cert_count, source,
               p.user.created_at.strftime('%d-%m-%Y')]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=col, value=value)
            cell.alignment = Alignment(vertical='center')
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F0F4FF',
                                        end_color='F0F4FF', fill_type='solid')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
        download_name='CIDI_Participants.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── DOWNLOAD IMPORT TEMPLATE ──────────────────────────────────────────────────
@admin.route('/import/template/participants')
@login_required
@admin_required
def download_participant_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Participants'

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='1A56DB',
                               end_color='1A56DB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center',
                             wrap_text=True)

    headers = [
        'full_name', 'email', 'password', 'gender',
        'place_of_birth', 'date_of_birth', 'city', 'address',
        'phone_number', 'education', 'working_status', 'skill_certified'
    ]
    notes = [
        'Full name of participant',
        'Valid email address (must be unique)',
        'Initial password for login',
        'Male or Female',
        'City of birth',
        'Format: DD-MM-YYYY (e.g. 15-01-2000)',
        'Current city of residence',
        'Full address',
        'Phone number (e.g. +62812...)',
        'SMA/SMK, D3, S1, S2, or S3',
        'employed or unemployed',
        'Leave empty if not yet certified. If certified, enter skill name exactly as in the system'
    ]
    examples = [
        'John Doe', 'john@email.com', 'Password123',
        'Male', 'Jakarta', '15-01-2000',
        'Bandung', 'Jl. Contoh No.1', '+62812345678',
        'S1', 'unemployed', 'IoT (Internet of Things)'
    ]

    for col, (h, note, ex) in enumerate(zip(headers, notes, examples), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        comment        = Comment(f"ℹ️ {note}", "CIDI Admin")
        cell.comment   = comment
        ex_cell        = ws.cell(row=2, column=col, value=ex)
        ex_cell.fill   = PatternFill(start_color='FFF3CD',
                                     end_color='FFF3CD', fill_type='solid')
        ex_cell.font   = Font(italic=True, color='856404')

    widths = [25, 30, 20, 10, 20, 18, 18, 30, 18, 12, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20

    # Valid skills list sheet
    ws3 = wb.create_sheet('Valid Skills')
    ws3.append(['Valid Skill Names for skill_certified column'])
    ws3.append(['IoT (Internet of Things)'])
    ws3.append(['Artificial Intelligence'])
    ws3.append(['Big Data Analytics'])
    ws3.append(['Cybersecurity'])
    ws3.append(['Cloud Computing'])
    ws3.append(['Robotics & Automation'])
    ws3.append(['3D Printing'])
    ws3.append(['Digital Manufacturing'])
    ws3['A1'].font = Font(bold=True, size=12, color='1A56DB')
    ws3.column_dimensions['A'].width = 35

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ['CIDI 4.0 - Participant Import Template Instructions'],
        [''],
        ['IMPORTANT RULES:'],
        ['1. Do NOT delete or rename the column headers in row 1'],
        ['2. Row 2 (yellow) is an EXAMPLE — replace with real data'],
        ['3. Start entering real data from row 3 onwards'],
        ['4. date_of_birth format must be: DD-MM-YYYY (e.g. 15-01-2000)'],
        ['5. gender must be exactly: Male or Female'],
        ['6. education must be one of: SMA/SMK, D3, S1, S2, S3'],
        ['7. working_status must be: employed or unemployed'],
        ['8. email must be unique — no duplicates allowed'],
        ['9. skill_certified: see the "Valid Skills" sheet for exact names'],
        ['10. If participant has no certificate yet, leave skill_certified empty'],
        ['11. All fields except skill_certified are required'],
        [''],
        ['After filling in the data, save the file and upload it in the Admin panel.'],
    ]
    for row_data in instructions:
        ws2.append(row_data)
    ws2.column_dimensions['A'].width = 70
    ws2['A1'].font = Font(bold=True, size=13, color='1A56DB')
    ws2['A3'].font = Font(bold=True, size=11)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
        download_name='CIDI_Participant_Import_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── IMPORT PARTICIPANTS FROM EXCEL ────────────────────────────────────────────
@admin.route('/import/participants', methods=['GET', 'POST'])
@login_required
@admin_required
def import_participants():
    if request.method == 'POST':
        import openpyxl
        file = request.files.get('excel_file')
        if not file or not file.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx Excel file.', 'danger')
            return redirect(url_for('admin.import_participants'))

        try:
            wb   = openpyxl.load_workbook(file)
            ws   = wb.active
            success_count = 0
            error_rows    = []

            for row_num, row in enumerate(
                ws.iter_rows(min_row=3, values_only=True), start=3
            ):
                if not any(row):
                    continue

                try:
                    # Read all 12 columns
                    full_name      = row[0]
                    email          = row[1]
                    password       = row[2]
                    gender         = row[3]
                    place_of_birth = row[4]
                    dob_str        = row[5]
                    city           = row[6]
                    address        = row[7]
                    phone_number   = row[8]
                    education      = row[9]
                    working_status = row[10]
                    skill_certified = row[11] if len(row) > 11 else None

                    # Validate required fields
                    if not full_name or not email or not password:
                        error_rows.append(
                            f"Row {row_num}: Missing required fields")
                        continue

                    # Check duplicate email
                    if User.query.filter_by(email=str(email)).first():
                        error_rows.append(
                            f"Row {row_num}: Email '{email}' already exists")
                        continue

                    # Parse date of birth
                    if isinstance(dob_str, datetime):
                        dob = dob_str.date()
                    else:
                        dob = datetime.strptime(
                            str(dob_str), '%d-%m-%Y').date()

                    # Create user account
                    new_user = User(
                        email       = str(email).strip(),
                        password    = generate_password_hash(str(password)),
                        role        = 'participant',
                        is_approved = True,
                        is_imported = True
                    )
                    db.session.add(new_user)
                    db.session.flush()

                    # Create participant profile
                    profile = ParticipantProfile(
                        user_id        = new_user.id,
                        full_name      = str(full_name).strip(),
                        gender         = str(gender).strip() if gender else 'Male',
                        place_of_birth = str(place_of_birth).strip() if place_of_birth else '-',
                        date_of_birth  = dob,
                        city           = str(city).strip() if city else None,
                        address        = str(address).strip() if address else '-',
                        phone_number   = str(phone_number).strip() if phone_number else '-',
                        education      = str(education).strip() if education else 'S1',
                        working_status = str(working_status).strip() if working_status else 'unemployed'
                    )
                    db.session.add(profile)
                    db.session.flush()

                    # Create certificate if skill_certified is provided
                    if skill_certified and str(skill_certified).strip():
                        # Find a matching completed program
                        skill_name = str(skill_certified).strip()
                        program = TrainingProgram.query.filter(
                            TrainingProgram.skill_category.ilike(
                                f'%{skill_name}%')
                        ).first()

                        # If no matching program found use first available
                        if not program:
                            program = TrainingProgram.query.first()

                        if program:
                            cert_no = 'CIDI-' + str(
                                uuid.uuid4()).upper()[:12]
                            cert = Certificate(
                                participant_id  = profile.id,
                                program_id      = program.id,
                                certificate_no  = cert_no,
                                issued_date     = date.today(),
                                skill_certified = skill_name
                            )
                            db.session.add(cert)

                    db.session.commit()
                    success_count += 1

                except Exception as e:
                    db.session.rollback()
                    error_rows.append(f"Row {row_num}: {str(e)}")

            if success_count > 0:
                flash(
                    f'Successfully imported {success_count} participant(s)!',
                    'success')
            if error_rows:
                for err in error_rows:
                    flash(err, 'warning')

        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')

        return redirect(url_for('admin.participants'))

    return render_template('admin/import_participants.html')

# ── APPROVE INDUSTRIES ────────────────────────────────────────────────────────
@admin.route('/industries')
@login_required
@admin_required
def industries():
    all_industries = User.query.filter_by(role='industry').all()
    return render_template('admin/industries.html', industries=all_industries)

@admin.route('/industries/approve/<int:id>')
@login_required
@admin_required
def approve_industry(id):
    user = User.query.get_or_404(id)
    user.is_approved = True
    db.session.commit()

    # Only send email if self-registered
    if not user.is_imported:
        industry_name = user.industry_profile.industry_name \
                        if user.industry_profile else 'Your organization'
        body = f"""
        <p>Dear <strong>{industry_name}</strong>,</p>
        <p>Your account on the CIDI 4.0 System has been <strong>approved</strong>.</p>
        <p>You can now log in and access:</p>
        <ul>
            <li>View the certified workforce database</li>
            <li>Submit workforce demand requests</li>
            <li>Track your demand request status</li>
        </ul>
        """
        sent = send_email(
            subject    = 'CIDI 4.0 — Your Account Has Been Approved!',
            recipients = [user.email],
            html_body  = email_template(
                title        = '✅ Account Approved!',
                greeting     = '',
                body_html    = body,
                button_text  = 'Login Now',
                button_url   = 'http://127.0.0.1:5000/auth/login'
            )
        )
        if sent:
            flash('Industry approved and notification email sent!', 'success')
        else:
            flash('Industry approved! (Email could not be sent — check MAIL settings in .env)', 'warning')
    else:
        flash('Industry account approved! (Imported account — no email sent)', 'success')

    return redirect(url_for('admin.industries'))

@admin.route('/industries/reject/<int:id>')
@login_required
@admin_required
def reject_industry(id):
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('Industry account rejected and removed.', 'success')
    return redirect(url_for('admin.industries'))

@admin.route('/industries/delete/<int:id>')
@login_required
@admin_required
def delete_industry(id):
    user = User.query.get_or_404(id)

    # Delete related records first
    if user.industry_profile:
        db.session.delete(user.industry_profile)
    WorkforceDemand.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    flash('Industry account deleted successfully.', 'success')
    return redirect(url_for('admin.industries'))

# ── APPROVE REGISTRATIONS ─────────────────────────────────────────────────────
@admin.route('/registrations')
@login_required
@admin_required
def registrations():
    all_regs = TrainingRegistration.query.order_by(
                   TrainingRegistration.registered_at.desc()).all()
    return render_template('admin/registrations.html', registrations=all_regs)

@admin.route('/registrations/approve/<int:id>')
@login_required
@admin_required
def approve_registration(id):
    reg = TrainingRegistration.query.get_or_404(id)
    reg.status = 'approved'
    db.session.commit()

    # Only send email if self-registered
    if not reg.participant.user.is_imported:
        body = f"""
        <p>Dear <strong>{reg.participant.full_name}</strong>,</p>
        <p>Your registration for the following training has been <strong>approved</strong>:</p>
        <div style="background:#eff6ff;border-left:4px solid #1a56db;
                    padding:1rem;margin:1rem 0;border-radius:4px;">
            <strong>{reg.program.title}</strong><br>
            <span style="color:#6b7280;">Skill: {reg.program.skill_category}</span>
        </div>
        <p>Please log in to view your training details.</p>
        """
        sent = send_email(
            subject    = f'CIDI 4.0 — Training Registration Approved: {reg.program.title}',
            recipients = [reg.participant.user.email],
            html_body  = email_template(
                title       = '✅ Registration Approved!',
                greeting    = '',
                body_html   = body,
                button_text = 'View My Training',
                button_url  = 'http://127.0.0.1:5000/auth/login'
            )
        )
        if sent:
            flash('Registration approved and notification email sent!', 'success')
        else:
            flash('Registration approved! (Email could not be sent — check MAIL settings in .env)', 'warning')
    else:
        flash('Registration approved! (Imported account — no email sent)', 'success')

    return redirect(url_for('admin.registrations'))

@admin.route('/registrations/reject/<int:id>')
@login_required
@admin_required
def reject_registration(id):
    reg = TrainingRegistration.query.get_or_404(id)
    reg.status = 'rejected'
    db.session.commit()
    flash('Registration rejected.', 'success')
    return redirect(url_for('admin.registrations'))

@admin.route('/registrations/complete/<int:id>')
@login_required
@admin_required
def complete_registration(id):
    reg = TrainingRegistration.query.get_or_404(id)
    reg.status = 'completed'
    db.session.commit()
    flash('Marked as completed!', 'success')
    return redirect(url_for('admin.registrations'))

# ── CERTIFICATES ──────────────────────────────────────────────────────────────
@admin.route('/certificates')
@login_required
@admin_required
def certificates():
    all_certs = Certificate.query.order_by(Certificate.issued_date.desc()).all()
    return render_template('admin/certificates.html', certificates=all_certs)

@admin.route('/certificates/issue/<int:registration_id>')
@login_required
@admin_required
def issue_certificate(registration_id):
    reg      = TrainingRegistration.query.get_or_404(registration_id)
    existing = Certificate.query.filter_by(
        participant_id=reg.participant_id,
        program_id=reg.program_id
    ).first()
    if existing:
        flash('Certificate already issued for this participant.', 'warning')
        return redirect(url_for('admin.registrations'))

    cert_no = 'CIDI-' + str(uuid.uuid4()).upper()[:12]
    cert = Certificate(
        participant_id  = reg.participant_id,
        program_id      = reg.program_id,
        certificate_no  = cert_no,
        issued_date     = date.today(),
        skill_certified = reg.program.skill_category
    )
    db.session.add(cert)
    reg.status = 'completed'
    db.session.commit()

    # Send certificate email if self-registered
    if not reg.participant.user.is_imported:
        body = f"""
        <p>Dear <strong>{reg.participant.full_name}</strong>,</p>
        <p>Congratulations! Your digital certificate has been issued.</p>
        <div style="background:#f0fdf4;border-left:4px solid #10b981;
                    padding:1rem;margin:1rem 0;border-radius:4px;">
            <strong>Certificate No:</strong> {cert_no}<br>
            <strong>Skill:</strong> {cert.skill_certified}<br>
            <strong>Program:</strong> {reg.program.title}<br>
            <strong>Issued:</strong> {cert.issued_date.strftime('%d %B %Y')}
        </div>
        <p>Log in to download your certificate PDF.</p>
        """
        send_email(
            subject    = 'CIDI 4.0 — Your Digital Certificate is Ready!',
            recipients = [reg.participant.user.email],
            html_body  = email_template(
                title       = '🎉 Certificate Issued!',
                greeting    = '',
                body_html   = body,
                button_text = 'Download Certificate',
                button_url  = 'http://127.0.0.1:5000/auth/login'
            )
        )

    flash(f'Certificate {cert_no} issued successfully!', 'success')
    return redirect(url_for('admin.registrations'))

# ── WORKFORCE DEMANDS ─────────────────────────────────────────────────────────
@admin.route('/demands')
@login_required
@admin_required
def demands():
    skill_filter  = request.args.get('skill', '')
    status_filter = request.args.get('status', '')
    query = WorkforceDemand.query
    if skill_filter:
        query = query.filter(
            WorkforceDemand.skill_required.ilike(f'%{skill_filter}%'))
    if status_filter:
        query = query.filter_by(status=status_filter)
    all_demands = query.order_by(WorkforceDemand.submitted_at.desc()).all()
    return render_template('admin/demands.html', demands=all_demands)

@admin.route('/demands/<int:id>')
@login_required
@admin_required
def view_demand(id):
    demand = WorkforceDemand.query.get_or_404(id)
    return render_template('admin/view_demand.html', demand=demand)

# ── ALL USERS ─────────────────────────────────────────────────────────────────
@admin.route('/users')
@login_required
@admin_required
def all_users():
    participants = User.query.filter_by(role='participant').order_by(
                       User.created_at.desc()).all()
    industries   = User.query.filter_by(role='industry').order_by(
                       User.created_at.desc()).all()
    return render_template('admin/all_users.html',
        participants=participants, industries=industries)

# ── CHANGE PASSWORD ───────────────────────────────────────────────────────────
@admin.route('/change-password', methods=['GET', 'POST'])
@login_required
@admin_required
def change_password():
    from werkzeug.security import check_password_hash
    if request.method == 'POST':
        current_pw = request.form.get('current_password')
        new_pw     = request.form.get('new_password')
        confirm_pw = request.form.get('confirm_password')
        if not check_password_hash(current_user.password, current_pw):
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('admin.change_password'))
        if new_pw != confirm_pw:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('admin.change_password'))
        current_user.password = generate_password_hash(new_pw)
        db.session.commit()
        flash('Password changed successfully!', 'success')
        return redirect(url_for('admin.dashboard'))
    return render_template('admin/change_password.html')


# ── AI TRAINING RECOMMENDATION ────────────────────────────────────────────────
@admin.route('/ai-recommendation')
@login_required
@admin_required
def ai_recommendation():
    pending_demands = WorkforceDemand.query.filter_by(
        status='pending'
    ).order_by(WorkforceDemand.submitted_at.desc()).all()
    all_demands = WorkforceDemand.query.order_by(
        WorkforceDemand.submitted_at.desc()
    ).limit(20).all()
    return render_template('admin/ai_recommendation.html',
        pending_demands = pending_demands,
        all_demands     = all_demands
    )

@admin.route('/ai-recommendation/analyze', methods=['POST'])
@login_required
@admin_required
def ai_analyze():
    demand_ids = request.form.getlist('demand_ids')
    if not demand_ids:
        flash('Please select at least one demand to analyze.', 'warning')
        return redirect(url_for('admin.ai_recommendation'))

    selected_demands = WorkforceDemand.query.filter(
        WorkforceDemand.id.in_(demand_ids)
    ).all()

    demands_text = ""
    for i, d in enumerate(selected_demands, 1):
        industry_name = d.user.industry_profile.industry_name \
                        if d.user.industry_profile else d.user.email
        demands_text += f"""
Demand #{i}:
- Industry: {industry_name}
- Skill Required: {d.skill_required}
- Job Position: {d.job_position or 'Not specified'}
- Machines/Technologies: {d.machines_used or 'Not specified'}
- Quantity Needed: {d.quantity} workers
- Experience Level: {d.experience_level or 'Not specified'}
- Minimum Education: {d.min_education or 'Not specified'}
- Work Location: {d.work_location or 'Not specified'}
- Urgency: {d.urgency_level or 'Not specified'}
- Description: {d.description or 'None'}
"""

    # Call Claude API from Python server
    import requests as req
    ai_result = None
    ai_error  = None

    try:
        response = req.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'Content-Type': 'application/json',
                'x-api-key': os.getenv('ANTHROPIC_API_KEY', ''),
                'anthropic-version': '2023-06-01'
            },
            json={
                'model': 'claude-opus-4-5',
                'max_tokens': 1000,
                'messages': [{
                    'role': 'user',
                    'content': f"""You are an expert training program designer for Industry 4.0 workforce development at the Center for Digital Industry 4.0 (CIDI 4.0) in Indonesia.

Based on the following workforce demand requests submitted by industries, recommend ONE specific training program that should be conducted.

WORKFORCE DEMANDS:
{demands_text}

IMPORTANT FORMATTING RULES:
- Do NOT use markdown formatting (no asterisks **, no pound signs #, no underscores)
- Use plain text only, with simple dashes (-) for lists
- Keep section headers in capital letters as shown below

Please provide your recommendation in this exact format:

RECOMMENDED TRAINING PROGRAM
==============================
Program Title: [specific title]
Skill Category: [one of: IoT (Internet of Things), Artificial Intelligence, Big Data Analytics, Cybersecurity, Cloud Computing, Robotics & Automation, 3D Printing, Digital Manufacturing]
Suggested Duration: [e.g. 2 weeks, 1 month - in total working days]
Suggested Capacity: [number of participants]

WHY THIS TRAINING:
[2-3 sentences explaining why this specific training is needed based on the demands]

TRAINING CONTENT OUTLINE:
[4-6 plain text dash-bullet points, NO bold formatting, of what should be covered]

DETAILED DAILY SCHEDULE:
[Break down the entire training duration into a day-by-day or week-by-week schedule. Use this exact format for each line:
Day 1-3: [topic covered those days]
Day 4-6: [topic covered those days]
Continue until the full duration is covered. Be specific about what is taught each block of days.]

TARGET PARTICIPANTS:
[Who should attend - education level, experience level]

URGENCY ASSESSMENT:
[Assessment of how urgent this training is based on the demand submissions, plain text only]"""
                }]
            },
            timeout=30
        )
        data = response.json()
        if 'content' in data and data['content']:
            ai_result = data['content'][0]['text']
        else:
            ai_error = f"API Error: {data.get('error', {}).get('message', 'Unknown error')}"
    except Exception as e:
        ai_error = str(e)

    return render_template('admin/ai_analyzing.html',
        demands_text     = demands_text,
        demand_ids       = demand_ids,
        selected_demands = selected_demands,
        ai_result        = ai_result,
        ai_error         = ai_error
    )

@admin.route('/ai-recommendation/create-program', methods=['POST'])
@login_required
@admin_required
def ai_create_program():
    title          = request.form.get('title')
    description    = request.form.get('description')
    skill_category = request.form.get('skill_category')
    start_date     = request.form.get('start_date')
    end_date       = request.form.get('end_date')
    capacity       = request.form.get('capacity')

    program = TrainingProgram(
        title          = title,
        description    = description,
        skill_category = skill_category,
        start_date     = date.fromisoformat(start_date),
        end_date       = date.fromisoformat(end_date),
        capacity       = int(capacity),
        status         = 'open'
    )
    db.session.add(program)
    db.session.commit()
    flash(f'Training program "{title}" created from AI recommendation!', 'success')
    return redirect(url_for('admin.programs'))
